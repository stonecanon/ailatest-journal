#!/usr/bin/env python3
"""Merge every list/ data source into journals.json + domestic_*.json.

主库 SCI (英文期刊为主):
    WoS Core (SCIE/SSCI/AHCI/ESCI) + JCR 2025 归属标记 + ESI 22 大类
  + 中科院 2025 大类分区（完整版 + 长江大学第二来源 中文大类）
  + ShowJCR JCR 2024 IF / Quartile / Rank
  + ShowJCR 中科院分区 2025 小类分区 + Top
  + ShowJCR 中科院新锐版 2026 中文刊名 + CN 号 + 语种 + 新锐分区
  + ShowJCR 国际期刊预警名单 2025
  + ShowJCR CCF 推荐 2026 (A/B/C)
  + 中国科协 2025 高质量科技期刊分级 (T1/T2/T3 + 领域)

国内分级 tab (单独 JSON):
    CCF-T 中文科技期刊 2025
    中国科协 高质量科技期刊分级（国内期刊）
    浙大 / 学校 A 国内学术期刊分级
    （CSSCI / 北大核心 PDF 为扫描件，需 OCR — 后续补）

Match priority: ISSN > eISSN > normalized title.
"""
from __future__ import annotations
import csv
import gzip
import json
import re
import shutil
import subprocess
import sys
from collections import Counter, defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

ROOT = Path(__file__).resolve().parent.parent
LIST_DIR = ROOT / 'list'
DATA_DIR = ROOT / 'data'
DATA_DIR.mkdir(exist_ok=True)

INDEX_FILES = {
    'SCIE': 'Science Citation Index Expanded (SCIE).csv',
    'SSCI': 'Social Sciences Citation Index (SSCI).csv',
    'AHCI': 'Arts & Humanities Citation Index (AHCI).csv',
    'ESCI': 'Emerging Sources Citation Index (ESCI).csv',
}
JCR_FILE    = LIST_DIR / 'JCR 2025.csv'
ESI_FILE    = LIST_DIR / 'ESI全部期刊列表.xlsx'
CAS_FILE    = LIST_DIR / '2025中科院分区表完整版（附2023vs2025对比版）.xlsx'
CJU_FILE    = LIST_DIR / '中国科学院2025年期刊大类划分（仅供参考用）第二来源-长江大学.xlsx'

SHOW_JCR    = LIST_DIR / 'ShowJCR_JCR_2024.csv'
SHOW_FQB    = LIST_DIR / 'ShowJCR_中科院分区_2025.csv'
SHOW_XR     = LIST_DIR / 'ShowJCR_中科院新锐版_2026.csv'
SHOW_CCF    = LIST_DIR / 'ShowJCR_CCF推荐_2026.csv'
SHOW_CCFT   = LIST_DIR / 'ShowJCR_CCF-T_2025.csv'
ABDC_CANDIDATES = [
    LIST_DIR / 'ABDC-JQL-2025-v1-260326.xlsx',
    LIST_DIR / 'ABDC Journal Quality List 2025.xlsx',
    LIST_DIR / 'ABDC_Journal_Quality_List_2025.xlsx',
    LIST_DIR / 'ABDC-JQL-2025.csv',
    LIST_DIR / 'ABDC Journal Quality List 2025.csv',
    LIST_DIR / 'ABDC-JQL-2022-v3-100523.xlsx',
    LIST_DIR / 'ABDC Journal Quality List 2022.xlsx',
    LIST_DIR / 'ABDC_Journal_Quality_List_2022.xlsx',
    LIST_DIR / 'ABDC-JQL-2022.csv',
    LIST_DIR / 'ABDC Journal Quality List 2022.csv',
]
ABS_CANDIDATES = [
    LIST_DIR / 'Academic Journal Guide 2024期刊目录.xlsx',
    LIST_DIR / 'Academic Journal Guide 2024.xlsx',
    LIST_DIR / 'AJG2024.xlsx',
    LIST_DIR / 'AJG_2024.xlsx',
    LIST_DIR / 'AJG2024.csv',
]
SCOPUS_FILE  = LIST_DIR / 'scopus_source_list.xlsx'
EI_FILE      = LIST_DIR / 'CPXSourceList_102025.xlsx'
OAJ_FILE     = LIST_DIR / 'oaj_journals.json'
DOAJ_FILE    = LIST_DIR / 'doaj_journals.csv'

CNKX_JSON    = DATA_DIR / 'cnkx_tiers.json'
CNKX_RECORDS = DATA_DIR / 'cnkx_records.json'
CNKX_DOMAINS = DATA_DIR / 'cnkx_domains_59.json'
ZJU_JSON     = DATA_DIR / 'zju_tiers.json'
SCHOOL_A_JSON= DATA_DIR / 'school_a_tiers.json'
CSSCI_CORE_JSON = ROOT / 'generated' / 'cssci_core.json'
CSSCI_EXT_JSON  = ROOT / 'generated' / 'cssci_ext.json'
PKU_CORE_JSON   = ROOT / 'generated' / 'pku_core.json'
CNKI_MAJOR_FILE = LIST_DIR / 'cnki_leaf_journals.csv'
CNKI_MAJOR_JSON = DATA_DIR / 'cnki_major_journals.json'


# ───────────────────────── utils ─────────────────────────

def norm_title(s: str) -> str:
    if not s: return ''
    s = s.upper()
    s = re.sub(r'[^A-Z0-9]+', '', s)
    return s

def clean_issn(s) -> str:
    if s is None: return ''
    s = str(s).strip().upper()
    m = re.search(r'\b(\d{4})-?(\d{3}[\dX])\b', s)
    return f"{m.group(1)}-{m.group(2)}" if m else ''

def split_issn_pair(s):
    """ShowJCR 常见 "1234-5678/2345-6789" 或 "1234-5678\n2345-6789" """
    if not s: return '', ''
    parts = re.split(r'[\s/；;,]+', str(s))
    issns = [clean_issn(p) for p in parts if clean_issn(p)]
    return (issns[0] if issns else '',
            issns[1] if len(issns) > 1 else '')

def first_existing(paths):
    for p in paths:
        if p.exists():
            return p
    return None

def pick_col(row, names):
    for name in names:
        if name in row and row[name] not in (None, ''):
            return row[name]
    lowered = {str(k).strip().lower(): k for k in row.keys()}
    for name in names:
        k = lowered.get(name.lower())
        if k is not None and row.get(k) not in (None, ''):
            return row.get(k)
    return ''


# ───────────────────────── WoS Core ─────────────────────────

def parse_wos_csv(path, index_name, store):
    added = 0
    with open(path, 'r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            title = (row.get('Journal title') or '').strip()
            if not title: continue
            issn = clean_issn(row.get('ISSN'))
            eissn = clean_issn(row.get('eISSN'))
            publisher = (row.get('Publisher name') or '').strip()
            addr = (row.get('Publisher address') or '').strip()
            langs = (row.get('Languages') or '').strip()
            cats = (row.get('Web of Science Categories') or '').strip()
            wos_cats = [c.strip() for c in cats.split('|') if c.strip()] if cats else []

            key = issn or eissn or norm_title(title)
            if not key: continue
            rec = store.get(key)
            if rec is None:
                rec = {
                    'name': title, 'issn': issn, 'eissn': eissn,
                    'publisher': publisher, 'address': addr, 'languages': langs,
                    'wos_categories': wos_cats, 'esi_category': '',
                    'abbr20': '', 'country': '', 'indices': [],
                }
                store[key] = rec
                added += 1
            else:
                if not rec['issn'] and issn: rec['issn'] = issn
                if not rec['eissn'] and eissn: rec['eissn'] = eissn
                if not rec['publisher'] and publisher: rec['publisher'] = publisher
                if not rec['address'] and addr: rec['address'] = addr
                if not rec['languages'] and langs: rec['languages'] = langs
                if not rec['wos_categories'] and wos_cats: rec['wos_categories'] = wos_cats
            if index_name not in rec['indices']:
                rec['indices'].append(index_name)
    return added


# ───────────────────────── JCR 2025 (Clarivate 自带表) ─────────────────────────

def parse_jcr(path, store, by_title):
    if not path.exists(): return 0
    hits = 0
    with open(path, 'r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            title = (row.get('Title') or '').strip()
            abbr20 = (row.get('Title20') or '').strip()
            country = (row.get('Country') or '').strip()
            if not title: continue
            flags = {k: (row.get(k) or '').strip() for k in ('SCIE','SSCI','AHCI','ESCI')}
            nt = norm_title(title)
            rec = by_title.get(nt)
            if rec is None:
                rec = {
                    'name': title, 'issn': '', 'eissn': '',
                    'publisher': '', 'address': '', 'languages': '',
                    'wos_categories': [], 'esi_category': '',
                    'abbr20': abbr20, 'country': country, 'indices': [],
                }
                store[nt] = rec
                by_title[nt] = rec
            if abbr20 and not rec.get('abbr20'): rec['abbr20'] = abbr20
            if country and not rec.get('country'): rec['country'] = country
            for idx, mark in flags.items():
                if mark.upper() == 'X' and idx not in rec['indices']:
                    rec['indices'].append(idx)
            hits += 1
    return hits


# ───────────────────────── ESI ─────────────────────────

def parse_esi(path, by_issn, by_title):
    if not openpyxl or not path.exists(): return 0
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    hits = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]: continue
        full = (row[0] or '').strip()
        cat  = (row[3] or '').strip() if len(row) > 3 else ''
        issn  = clean_issn(row[4] if len(row) > 4 and row[4] else '')
        eissn = clean_issn(row[5] if len(row) > 5 and row[5] else '')
        nt = norm_title(full)
        rec = (issn and by_issn.get(issn)) or (eissn and by_issn.get(eissn)) or by_title.get(nt)
        if rec and cat and not rec.get('esi_category'):
            rec['esi_category'] = cat
            hits += 1
    return hits


# ───────────────────────── 中科院 2025 完整版 ─────────────────────────

def parse_cas(path, by_title):
    """完整版表头: 期刊名称 | 2025分区 | 2023分区 | Top | Open Access"""
    if not openpyxl or not path.exists(): return (0, 0)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['2023 vs 2025分区对比'] if '2023 vs 2025分区对比' in wb.sheetnames else wb.active
    hits = miss = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]: continue
        title = str(row[0]).strip()
        z25 = row[1] if len(row) > 1 else None
        z23 = row[2] if len(row) > 2 else None
        top = row[3] if len(row) > 3 else None
        oa  = row[4] if len(row) > 4 else None
        nt = norm_title(title)
        rec = by_title.get(nt)
        if rec is None:
            miss += 1
            continue
        try: rec['cas_zone'] = int(z25) if z25 is not None and str(z25).strip() else None
        except (TypeError, ValueError): rec['cas_zone'] = None
        try: rec['cas_zone_2023'] = int(z23) if z23 is not None and str(z23).strip() else None
        except (TypeError, ValueError): rec['cas_zone_2023'] = None
        rec['cas_top'] = (str(top).strip() == '是') if top is not None else False
        rec['cas_oa']  = (str(oa).strip() == '是')  if oa is not None else False
        hits += 1
    return hits, miss


# ───────────────────────── 长江大学 中文大类 ─────────────────────────

def parse_changjiang(path, by_issn, by_title):
    """表头: 刊名 | ISSN | EISSN | 数据库 | 大类 | 大类(英文) | 出版社"""
    if not openpyxl or not path.exists(): return 0
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    hits = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]: continue
        title = str(row[0]).strip()
        issn  = clean_issn(row[1] if len(row) > 1 else '')
        eissn = clean_issn(row[2] if len(row) > 2 else '')
        cat_cn = str(row[4]).strip() if len(row) > 4 and row[4] else ''
        nt = norm_title(title)
        rec = (issn and by_issn.get(issn)) or (eissn and by_issn.get(eissn)) or by_title.get(nt)
        if rec and cat_cn:
            rec['cas_major_cn'] = cat_cn
            hits += 1
    return hits


# ───────────────────────── ShowJCR warning ─────────────────────────

def parse_warning(path, by_title, by_issn):
    if not path.exists(): return 0
    hits = 0
    with open(path, 'r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            title = (row.get('Journal') or row.get('journal') or '').strip()
            reason = (row.get('预警原因(2025)') or row.get('预警原因') or '').strip()
            if not title: continue
            nt = norm_title(title)
            rec = by_title.get(nt)
            if rec is None: continue
            rec['warning'] = reason or True
            hits += 1
    return hits


WARNING_XLSX = LIST_DIR / '国际期刊预警名单_2020-2025.xlsx'
def parse_warning_xlsx(path, by_title, by_issn):
    """Parse the comprehensive CAS warning list xlsx (2020-2025).
    '全部名单' sheet columns: 序号, 年份, 学科, 期刊, ISSN/EISSN, 预警级别, 预警原因, 发布信息, 页面链接
    Stores warning as object {year, level, reason, subject} — supports multiple years.
    """
    if not path.exists():
        print(f'  WARNING: {path} not found')
        return 0
    try:
        import openpyxl
    except ImportError:
        print('  WARNING: openpyxl not available, skipping warning xlsx')
        return 0
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if '全部名单' not in wb.sheetnames:
        print(f'  WARNING: no "全部名单" sheet in {path}')
        return 0
    ws = wb['全部名单']
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0
    headers = rows[0]
    # find column indices
    try:
        idx_year   = headers.index('年份')
        idx_subj   = headers.index('学科')
        idx_jrnl   = headers.index('期刊')
        idx_issn   = headers.index('ISSN/EISSN')
        idx_level  = headers.index('预警级别')
        idx_reason = headers.index('预警原因')
    except ValueError as e:
        print(f'  WARNING: column mismatch in "全部名单": {e}')
        return 0
    hits = 0
    for row in rows[1:]:
        if not row or not row[idx_jrnl]:
            continue
        title = str(row[idx_jrnl]).strip()
        issn  = str(row[idx_issn]).strip() if row[idx_issn] else ''
        year  = int(row[idx_year]) if row[idx_year] else None
        level = str(row[idx_level]).strip() if row[idx_level] else None
        subject = str(row[idx_subj]).strip() if row[idx_subj] else None
        reason = str(row[idx_reason]).strip() if row[idx_reason] else None
        if not title:
            continue
        nt = norm_title(title)
        rec = by_title.get(nt)
        if rec is None and issn and issn != '-':
            rec = by_issn.get(issn)
        if rec is None:
            continue
        wobj = {'year': year, 'level': level, 'reason': reason, 'subject': subject}
        existing = rec.get('warning')
        if isinstance(existing, list):
            existing.append(wobj)
        elif isinstance(existing, dict):
            rec['warning'] = [existing, wobj]
        else:
            rec['warning'] = wobj
        hits += 1
    wb.close()
    return hits


# ───────────────────────── ShowJCR JCR 2024 (IF) ─────────────────────────

def parse_showjcr_if(path, by_title, by_issn):
    """表头: Journal, ISSN, eISSN, Category, IF(2024), IF Quartile(2024), IF Rank(2024)"""
    if not path.exists(): return 0
    hits = 0
    # Clean JCR category suffix like (SCIE)/(SSCI)/(AHCI)/(ESCI)
    _cat_pat = re.compile(r'\s*\((?:SCIE|SSCI|AHCI|ESCI)\)\s*')
    with open(path, 'r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            title = (row.get('Journal') or '').strip()
            if not title: continue
            issn = clean_issn(row.get('ISSN'))
            eissn = clean_issn(row.get('eISSN'))
            nt = norm_title(title)
            rec = (issn and by_issn.get(issn)) or (eissn and by_issn.get(eissn)) or by_title.get(nt)
            if rec is None: continue
            if_val = (row.get('IF(2024)') or '').strip()
            if if_val:
                try: rec['if_2024'] = float(if_val)
                except ValueError: pass
            q = (row.get('IF Quartile(2024)') or '').strip()
            if q: rec['if_quartile'] = q
            rk = (row.get('IF Rank(2024)') or '').strip()
            if rk: rec['if_rank'] = rk
            # Save JCR category (may contain multiple categories separated by semicolons)
            cat_raw = (row.get('Category') or '').strip()
            if cat_raw:
                # Clean index suffix from each category
                cats = [_cat_pat.sub('', c).strip() for c in re.split(r'[;；]', cat_raw) if c.strip()]
                if cats:
                    rec['jcr_cat'] = cats[0]  # primary category
                    if len(cats) > 1:
                        rec['jcr_cats'] = cats  # all categories
            hits += 1
    return hits


# ───────────────────────── ShowJCR FQB 2025 (小类分区) ─────────────────────────

def parse_showjcr_fqb(path, by_title, by_issn):
    """表头含: Journal, ISSN/EISSN, Review, OAJ, Open Access, WoS, 标注,
       大类, 大类分区, Top, 小类1..小类6 (每个格式 '名称 分区')"""
    if not path.exists(): return 0
    hits = 0
    with open(path, 'r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            title = (row.get('Journal') or '').strip()
            if not title: continue
            issn_raw = row.get('ISSN/EISSN') or row.get('ISSN') or ''
            issn, eissn = split_issn_pair(issn_raw)
            nt = norm_title(title)
            rec = (issn and by_issn.get(issn)) or (eissn and by_issn.get(eissn)) or by_title.get(nt)
            if rec is None: continue

            major = (row.get('大类') or '').strip()
            major_zone = (row.get('大类分区') or '').strip()
            is_top = (row.get('Top') or '').strip() == '是'

            if major: rec.setdefault('cas_major_cat', major)
            if major_zone:
                mm = re.search(r'\d', major_zone)
                if mm:
                    try: rec['cas_major_zone'] = int(mm.group(0))
                    except ValueError: pass
            if is_top and not rec.get('cas_top'):
                rec['cas_top'] = True

            subs = []
            for i in range(1, 7):
                nm = (row.get(f'小类{i}') or '').strip()
                if not nm: continue
                zn_raw = (row.get(f'小类{i}分区') or '').strip()
                zn = None
                if zn_raw:
                    mz = re.search(r'\d', zn_raw)
                    if mz:
                        try: zn = int(mz.group(0))
                        except ValueError: pass
                # legacy fallback: name + zone packed in one cell
                if zn is None:
                    m = re.match(r'^(.+?)\s+(\d)\s*区?\s*$', nm)
                    if m:
                        nm = m.group(1).strip()
                        zn = int(m.group(2))
                subs.append({'name': nm, 'zone': zn})
            if subs: rec['cas_sub_cats'] = subs
            hits += 1
    return hits


# ───────────────────────── ShowJCR XR 2026 (中文刊名/新锐版) ─────────────────────────

def parse_showjcr_xr(path, by_title, by_issn):
    """表头含: Journal, 年份, 预警标记, 刊名, 中文刊名, CN, ISSN, EISSN,
       出版机构, 语种, 期刊类型, 数据库,
       大类(英文/中文) 1..2 含新锐分区, Top, 小类1..6(英文/中文/新锐分区)"""
    if not path.exists(): return 0
    hits = 0

    def norm_zone(v):
        """'2 区' → '2', '—' → '', None/'' → ''"""
        if not v: return ''
        s = str(v).strip().replace(' ', '')
        if not s or s in ('—', '-', '－'): return ''
        m = re.match(r'^([1-4])', s)
        return m.group(1) if m else ''

    def norm_top(v):
        if not v: return False
        s = str(v).strip()
        return s in ('是', 'Y', 'YES', 'TRUE', '1', 'Top', 'TOP')

    with open(path, 'r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or []
        # 查找主要字段的实际 key
        def find(col_substr):
            for fn in fieldnames:
                if col_substr in fn: return fn
            return None

        col_cname   = find('中文刊名')
        col_cn      = find('CN')
        col_lang    = find('语种')
        col_issn    = 'ISSN' if 'ISSN' in fieldnames else find('ISSN')
        col_eissn   = 'EISSN' if 'EISSN' in fieldnames else find('EISSN')
        col_cat_cn  = None
        col_cat_en  = None
        col_cat2_cn = None
        col_cat2_en = None
        for fn in fieldnames:
            if '大类' in fn and '中文' in fn and '2' not in fn and not col_cat_cn:
                col_cat_cn = fn
            elif '大类' in fn and '英文' in fn and '2' not in fn and not col_cat_en:
                col_cat_en = fn
            elif '大类2' in fn and '中文' in fn and not col_cat2_cn:
                col_cat2_cn = fn
            elif '大类2' in fn and '英文' in fn and not col_cat2_en:
                col_cat2_en = fn

        # 新锐版分区列
        col_zone1   = '大类新锐分区' if '大类新锐分区' in fieldnames else None
        col_top1    = 'Top' if 'Top' in fieldnames else None
        col_zone2   = '大类2新锐分区' if '大类2新锐分区' in fieldnames else None
        col_top2    = '大类2Top' if '大类2Top' in fieldnames else None
        # 小类 1-6 (中文名 + 分区)
        sub_cols = []
        for n in range(1, 7):
            cn = f'小类{n}中文名'
            zn = f'小类{n}新锐分区'
            if cn in fieldnames and zn in fieldnames:
                sub_cols.append((cn, zn))

        for row in reader:
            title = (row.get('Journal') or row.get('刊名') or '').strip()
            if not title: continue
            issn  = clean_issn(row.get(col_issn, '')) if col_issn else ''
            eissn = clean_issn(row.get(col_eissn, '')) if col_eissn else ''
            nt = norm_title(title)
            rec = (issn and by_issn.get(issn)) or (eissn and by_issn.get(eissn)) or by_title.get(nt)
            if rec is None: continue
            if col_cname:
                cname = (row.get(col_cname) or '').strip()
                if cname and cname != title:
                    rec['cn_name'] = cname
            if col_cn:
                cn = (row.get(col_cn) or '').strip()
                if cn: rec['cn_code'] = cn
            if col_lang:
                lang = (row.get(col_lang) or '').strip()
                if lang: rec['language_cn'] = lang
            if col_cat_cn and not rec.get('cas_major_cn'):
                v = (row.get(col_cat_cn) or '').strip()
                if v: rec['cas_major_cn'] = v.split()[0] if v else ''

            # 新锐分区 → cas_xr
            xr = {}
            z1 = norm_zone(row.get(col_zone1)) if col_zone1 else ''
            if z1:
                xr['zone'] = z1
                if col_top1: xr['top'] = norm_top(row.get(col_top1))
            z2 = norm_zone(row.get(col_zone2)) if col_zone2 else ''
            if z2:
                xr['zone2'] = z2
                if col_top2: xr['top2'] = norm_top(row.get(col_top2))
            # 把大类挂到 xr 上，前端展示用
            if col_cat_cn:
                v = (row.get(col_cat_cn) or '').strip()
                if v: xr['major_cn'] = v.split()[0] if v else ''
            if col_cat_en:
                v = (row.get(col_cat_en) or '').strip()
                if v: xr['major_en'] = v
            if col_cat2_cn:
                v = (row.get(col_cat2_cn) or '').strip()
                if v: xr['major2_cn'] = v.split()[0] if v else ''
            if col_cat2_en:
                v = (row.get(col_cat2_en) or '').strip()
                if v: xr['major2_en'] = v
            subs = []
            for cn_col, zn_col in sub_cols:
                name = (row.get(cn_col) or '').strip()
                z = norm_zone(row.get(zn_col))
                if name and z:
                    subs.append({'name': name, 'zone': z})
            if subs:
                xr['sub'] = subs
            if xr:
                rec['cas_xr'] = xr

            hits += 1
    return hits


# ───────────────────────── ShowJCR CCF 2026 (英文期刊，merge 回主库) ─────────────────────────

def parse_showjcr_ccf(path, by_title, by_issn):
    """表头: 刊物名称, Journal, 年份, 出版社, 网址, 领域, CCF推荐类别, CCF推荐类型"""
    if not path.exists(): return 0
    hits = 0
    with open(path, 'r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            title = (row.get('Journal') or row.get('刊物名称') or '').strip()
            if not title: continue
            nt = norm_title(title)
            rec = by_title.get(nt)
            if rec is None: continue
            ccf = (row.get('CCF推荐类型') or row.get('CCF推荐类型 A/B/C') or '').strip()
            area = (row.get('领域') or row.get('CCF推荐类别') or '').strip()
            if ccf: rec['ccf'] = ccf
            if area: rec['ccf_area'] = area
            hits += 1
    return hits


# ───────────────────────── ABDC Journal Quality List ─────────────────────────

def parse_abdc(path, by_title, by_issn, store=None):
    """ABDC Journal Quality List: A*, A, B, C.

    Expected official/common headers:
      Journal Title | Publisher | ISSN | ISSN Online | Year Inception | FoR | <year> rating
    The parser is deliberately tolerant because downloaded workbooks sometimes vary
    in spacing/casing.
    """
    if not path: return 0
    source_year = '2025' if '2025' in path.name else '2022' if '2022' in path.name else ''

    def apply_row(row):
        title = str(pick_col(row, ['Journal Title', 'Title', 'Journal', 'Journal title']) or '').strip()
        if not title:
            return 0
        issn = clean_issn(pick_col(row, ['ISSN', 'Print ISSN', 'ISSN Print']))
        eissn = clean_issn(pick_col(row, ['ISSN Online', 'ISSNOnline', 'Online ISSN', 'eISSN', 'EISSN', 'ISSN_Online']))
        rating = str(pick_col(row, ['2025 rating', '2022 rating', 'Rating', 'ABDC Rating', 'ABDC 2025', 'ABDC 2022', 'Rank']) or '').strip().upper()
        rating = rating.replace('A STAR', 'A*').replace('A-STAR', 'A*').replace('A* ', 'A*')
        if rating not in {'A*', 'A', 'B', 'C'}:
            return 0
        nt = norm_title(title)
        rec = (issn and by_issn.get(issn)) or (eissn and by_issn.get(eissn)) or by_title.get(nt)
        field = str(pick_col(row, ['FoR', 'FOR', 'Field of Research', 'Field', 'FoR code']) or '').strip()
        publisher = str(pick_col(row, ['Publisher']) or '').strip()
        if rec is None:
            if store is None:
                return 0
            key = 'abdc:' + (issn or eissn or nt)
            rec = store.get(key)
            if rec is None:
                rec = {
                    'name': title, 'issn': issn, 'eissn': eissn,
                    'publisher': publisher, 'address': '', 'languages': '',
                    'wos_categories': [], 'esi_category': '',
                    'abbr20': '', 'country': '', 'indices': [],
                    'abdc_only': True,
                }
                store[key] = rec
                by_title.setdefault(nt, rec)
                for k in (issn, eissn):
                    if k: by_issn.setdefault(k, rec)
        rec['abdc'] = {
            'rating': rating,
            'field': field,
            'source': f'ABDC Journal Quality List {source_year}'.strip(),
        }
        if publisher and not rec.get('publisher'):
            rec['publisher'] = publisher
        return 1

    hits = 0
    if path.suffix.lower() == '.csv':
        with open(path, 'r', encoding='utf-8-sig', newline='') as f:
            for row in csv.DictReader(f):
                hits += apply_row(row)
        return hits

    if not openpyxl:
        return 0
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = None
    for raw in rows:
        vals = [str(v).strip() if v is not None else '' for v in raw]
        joined = ' '.join(vals).lower()
        if 'journal' in joined and ('rating' in joined or 'issn' in joined):
            headers = vals
            break
    if not headers:
        return 0
    for h in headers:
        m = re.search(r'\b(20\d{2})\s+rating\b', h, re.I)
        if m:
            source_year = m.group(1)
            break
    for raw in rows:
        row = {headers[i]: raw[i] if i < len(raw) else '' for i in range(len(headers))}
        hits += apply_row(row)
    return hits


# ───────────────────────── ABS / AJG Academic Journal Guide ─────────────────────────

def parse_abs(path, by_title, by_issn, store=None):
    """Chartered ABS Academic Journal Guide (AJG): 4*, 4, 3, 2, 1.

    Source: https://charteredabs.org/academic-journal-guide/academic-journal-guide-2024
    Expected columns: ID | Field | Title (or Ttitle) | AJG_2024
    Match strategy: title-only (the source xlsx ships no ISSN).
    """
    if not path or not openpyxl: return 0
    source_year = '2024'
    m = re.search(r'(20\d{2})', path.name)
    if m: source_year = m.group(1)

    valid = {'4*', '4', '3', '2', '1'}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers = None
    rows_iter = ws.iter_rows(values_only=True)
    for raw in rows_iter:
        vals = [str(v).strip() if v is not None else '' for v in raw]
        joined = ' '.join(vals).lower()
        if ('title' in joined or 'ttitle' in joined) and ('ajg' in joined or 'rating' in joined or 'rank' in joined):
            headers = vals
            break
    if not headers:
        return 0

    def find_idx(keys):
        keys = [k.lower() for k in keys]
        for i, h in enumerate(headers):
            hl = h.strip().lower()
            if hl in keys: return i
        return -1

    i_title = find_idx(['title', 'ttitle', 'journal title', 'journal'])
    i_field = find_idx(['field', 'subject', 'area'])
    i_rate  = find_idx(['ajg_2024', 'ajg 2024', 'ajg2024', 'rating', 'rank', '2024 rating'])
    if i_title < 0 or i_rate < 0:
        return 0

    hits = 0
    standalone = 0
    for raw in rows_iter:
        def cell(i):
            if i < 0 or i >= len(raw): return ''
            v = raw[i]
            return '' if v is None else str(v).strip()
        title = cell(i_title)
        if not title: continue
        rating = cell(i_rate).replace('★', '*').replace('（', '').replace('）', '').strip()
        if rating not in valid:
            try:
                f = float(rating)
                if f.is_integer():
                    rating = str(int(f))
            except Exception:
                pass
            if rating not in valid:
                continue
        nt = norm_title(title)
        rec = by_title.get(nt)
        abs_payload = {
            'rating': rating,
            'field': cell(i_field),
            'source': f'Chartered ABS AJG {source_year}',
        }
        if rec is not None:
            rec['abs'] = abs_payload
            hits += 1
        elif store is not None:
            # ABS-only record (not in WoS Core) — usually management/IB titles
            key = 'abs:' + nt
            if key in store:
                store[key]['abs'] = abs_payload
            else:
                store[key] = {
                    'name': title, 'issn': '', 'eissn': '',
                    'wos_categories': [], 'esi_category': '',
                    'abbr20': '', 'country': '', 'indices': [],
                    'abs': abs_payload,
                    'abs_only': True,
                }
            by_title[nt] = store[key]
            standalone += 1
    if standalone:
        print(f'  ABS standalone (not in WoS): +{standalone}')
    return hits + standalone


# ───────────────────────── Scopus Source List ─────────────────────────

def parse_scopus(path, by_title, by_issn, store=None):
    """Scopus Source List (Mar. 2026).

    Match priority: ISSN > EISSN > normalized title. Journal-only rows are used.
    Active matched rows get a visible Scopus badge; inactive rows keep metadata for drawer notes.
    Active journal rows not found in WoS are added as Scopus-only records so the list source is not lost.
    """
    if not path.exists() or not openpyxl:
        return 0, 0, 0
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # auto-detect sheet containing "Scopus Sources"
    sheet_name = next((s for s in wb.sheetnames if 'Scopus Sources' in s), None)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(v).strip() if v is not None else '' for v in next(rows)]
    col = {h: i for i, h in enumerate(headers) if h}

    def cell(raw, name):
        i = col.get(name)
        if i is None or i >= len(raw): return ''
        v = raw[i]
        return '' if v is None else str(v).strip()

    top_cols = []
    for label, token in [
        ('Life Sciences', 'Life Sciences'),
        ('Social Sciences', 'Social Sciences'),
        ('Physical Sciences', 'Physical Sciences'),
        ('Health Sciences', 'Health Sciences'),
    ]:
        for h, i in col.items():
            if 'Top level' in h and token in h:
                top_cols.append((label, i)); break

    matched = standalone = inactive = 0
    for raw in rows:
        title = cell(raw, 'Source Title')
        if not title: continue
        if cell(raw, 'Source Type').lower() != 'journal':
            continue
        issn = clean_issn(cell(raw, 'ISSN'))
        eissn = clean_issn(cell(raw, 'EISSN'))
        nt = norm_title(title)
        status = cell(raw, 'Active or Inactive')
        active = status.lower() != 'inactive'
        asjc = [s.strip() for s in re.split(r'[;,]+', cell(raw, 'All Science Journal Classification Codes (ASJC)')) if s and s.strip()]
        asjc_top = []
        for label, i in top_cols:
            v = raw[i] if i < len(raw) else None
            if v not in (None, '') and label not in asjc_top:
                asjc_top.append(label)
        payload = {
            'id': cell(raw, 'Sourcerecord ID'),
            'active': active,
            'coverage': cell(raw, 'Coverage'),
            'asjc': asjc,
            'asjc_top': asjc_top,
            'source': 'Scopus Source List (auto-updated)',
        }
        rec = (issn and by_issn.get(issn)) or (eissn and by_issn.get(eissn)) or by_title.get(nt)
        if rec is not None:
            rec['scopus'] = payload
            if not rec.get('publisher') and cell(raw, 'Publisher'):
                rec['publisher'] = cell(raw, 'Publisher')
            matched += 1
            if not active: inactive += 1
        elif store is not None and active:
            key = 'scopus:' + (issn or eissn or nt)
            rec = store.get(key)
            if rec is None:
                rec = {
                    'name': title, 'issn': issn, 'eissn': eissn,
                    'publisher': cell(raw, 'Publisher'), 'address': '', 'languages': cell(raw, 'Article Language in Source (Three-Letter ISO Language Codes)'),
                    'wos_categories': [], 'esi_category': '',
                    'abbr20': '', 'country': '', 'indices': [],
                    'scopus_only': True,
                }
                store[key] = rec
                by_title.setdefault(nt, rec)
                for k in (issn, eissn):
                    if k: by_issn.setdefault(k, rec)
            rec['scopus'] = payload
            standalone += 1
    return matched, standalone, inactive


# ───────────────────────── EI Compendex Source List ─────────────────────────

def parse_ei_compendex(path, by_title, by_issn, store=None):
    """EI/Compendex Source List (Oct. 2025) journal rows + Chinese journals sheet."""
    if not path.exists() or not openpyxl:
        return 0, 0, 0
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    matched = standalone = chinese_hits = 0

    def apply(title, issn, eissn, publisher='', country='', language='', subjects=None, status='', cn_name=''):
        nonlocal matched, standalone, chinese_hits
        if not title: return
        issn = clean_issn(issn); eissn = clean_issn(eissn)
        nt = norm_title(title)
        subjects = [s for s in (subjects or []) if s and str(s).strip() not in {'-', '—'}]
        subjects = [str(s).strip() for s in subjects]
        rec = (issn and by_issn.get(issn)) or (eissn and by_issn.get(eissn)) or by_title.get(nt)
        if rec is None and store is not None:
            key = 'ei:' + (issn or eissn or nt)
            rec = store.get(key)
            if rec is None:
                rec = {
                    'name': title, 'issn': issn, 'eissn': eissn,
                    'publisher': publisher or '', 'address': '', 'languages': language or '',
                    'wos_categories': [], 'esi_category': '',
                    'abbr20': '', 'country': country or '', 'indices': [],
                    'ei_only': True,
                }
                store[key] = rec
                by_title.setdefault(nt, rec)
                for k in (issn, eissn):
                    if k: by_issn.setdefault(k, rec)
            standalone += 1
        elif rec is not None:
            matched += 1
        if rec is None:
            return
        if 'EI' not in rec.setdefault('indices', []):
            rec['indices'].append('EI')
        if subjects:
            cur = rec.setdefault('ei_subjects', [])
            for s in subjects:
                if s not in cur: cur.append(s)
        if publisher and not rec.get('publisher'): rec['publisher'] = publisher
        if country and not rec.get('country'): rec['country'] = country
        if language and not rec.get('languages'): rec['languages'] = language
        if status: rec['ei_status'] = status
        if cn_name:
            rec['cn_name'] = cn_name
            chinese_hits += 1

    if 'SERIALS' in wb.sheetnames:
        ws = wb['SERIALS']
        for raw in ws.iter_rows(min_row=3, values_only=True):
            if not raw or not raw[0]: continue
            stype = str(raw[1] or '').strip().lower() if len(raw) > 1 else ''
            if stype != 'journal': continue
            apply(raw[0], raw[2] if len(raw)>2 else '', raw[3] if len(raw)>3 else '',
                  publisher=str(raw[4] or '').strip() if len(raw)>4 else '',
                  country=str(raw[5] or '').strip() if len(raw)>5 else '',
                  language=str(raw[6] or '').strip() if len(raw)>6 else '',
                  subjects=list(raw[7:15]) if len(raw)>7 else [],
                  status='Compendex Source List Oct. 2025')

    if 'CHINESE JRS on SERIALS LIST' in wb.sheetnames:
        ws = wb['CHINESE JRS on SERIALS LIST']
        for raw in ws.iter_rows(min_row=3, values_only=True):
            if not raw or not raw[4]: continue
            apply(raw[4], raw[0] if len(raw)>0 else '', raw[1] if len(raw)>1 else '',
                  language=str(raw[5] or '').strip() if len(raw)>5 else '',
                  status=str(raw[6] or '').strip() if len(raw)>6 else '',
                  cn_name=str(raw[2] or '').strip() if len(raw)>2 else '')
    return matched, standalone, chinese_hits


# ───────────────────────── flagship Nature / Science / Cell ─────────────────────────

def infer_flagship(title, publisher=''):
    n = norm_title(title)
    pub = (publisher or '').upper()
    if n == 'NATURE': return 'nature_main'
    if n == 'SCIENCE': return 'science_main'
    if n == 'CELL': return 'cell_main'
    t = (title or '').strip().upper()
    if t.startswith('NATURE ') or t.startswith('NATURE-') or t.startswith('NATURE REVIEWS'):
        return 'nature_sub'
    if t in {'SCIENCE ADVANCES', 'SCIENCE ROBOTICS', 'SCIENCE IMMUNOLOGY', 'SCIENCE SIGNALING', 'SCIENCE TRANSLATIONAL MEDICINE'}:
        return 'science_sub'
    if t.startswith('CELL ') or t.startswith('CELL REPORTS') or t in {'CELL GENOMICS', 'CELL METABOLISM', 'CELL STEM CELL', 'CELL SYSTEMS', 'CELL HOST & MICROBE'}:
        return 'cell_sub'
    if 'NATURE PORTFOLIO' in pub and t.startswith('NATURE'):
        return 'nature_sub'
    return ''


# ───────────────────────── 中国科协 — merge 回主库 (按 ISSN) ─────────────────────────

def merge_cnkx_to_main(by_issn, by_title):
    if not CNKX_RECORDS.exists(): return 0
    data = json.loads(CNKX_RECORDS.read_text(encoding='utf-8'))
    hits = 0
    for r in data:
        issn = r.get('issn')
        nt = norm_title(r.get('name') or '')
        rec = (issn and by_issn.get(issn)) or by_title.get(nt)
        if rec is None: continue
        cnkx = rec.setdefault('cnkx', [])
        entry = {'domain': r.get('domain'), 'tier': r.get('tier') or None}
        if entry not in cnkx:
            cnkx.append(entry)
        hits += 1
    return hits


# ───────────────────────── OAJ 全球开放获取期刊索引 ─────────────────────────

def parse_oaj(path, by_title, by_issn, store=None):
    """读取 OAJ 开放获取期刊索引 JSON，匹配到现有记录或新增."""
    if not path.exists():
        return 0, 0
    with open(path, 'r', encoding='utf-8') as f:
        oaj_list = json.load(f)
    matched = standalone = 0
    for item in oaj_list:
        title = item.get('title') or ''
        issn = clean_issn(item.get('issn'))
        eissn = clean_issn(item.get('eissn'))
        nt = norm_title(title)
        payload = {
            'partition': item.get('partition'),   # e.g. "1区"
            'position': item.get('positioning'),   # e.g. "Frontier Science Journal"
            'oa_type': item.get('oa_type'),        # e.g. "Gold OA"
        }
        rec = (issn and by_issn.get(issn)) or (eissn and by_issn.get(eissn)) or by_title.get(nt)
        if rec is not None:
            rec['oaj'] = payload
            matched += 1
        elif store is not None:
            key = 'oaj:' + (issn or eissn or nt)
            rec = store.get(key)
            if rec is None:
                rec = {
                    'name': title, 'issn': issn, 'eissn': eissn,
                    'publisher': item.get('publisher', ''),
                    'country': item.get('country', ''),
                    'abbr20': '', 'indices': [], 'wos_categories': [], 'esi_category': '',
                    'oaj_only': True,
                }
                store[key] = rec
                by_title.setdefault(nt, rec)
                for k in (issn, eissn):
                    if k: by_issn.setdefault(k, rec)
            rec['oaj'] = payload
            standalone += 1
    return matched, standalone


# ───────────────────────── DOAJ Directory of Open Access Journals ─────────────────────────

def parse_doaj(path, by_title, by_issn, store=None):
    """DOAJ Journal CSV.

    Official public CSV: https://doaj.org/csv
    Match priority: print ISSN > online ISSN > normalized title. DOAJ-only journals
    are added because many fully open-access journals are outside WoS/EI/Scopus.
    """
    if not path.exists():
        return 0, 0
    matched = standalone = 0
    with open(path, 'r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            title = (row.get('Journal title') or '').strip()
            if not title:
                continue
            issn = clean_issn(row.get('Journal ISSN (print version)'))
            eissn = clean_issn(row.get('Journal EISSN (online version)'))
            nt = norm_title(title)
            payload = {
                'u': (row.get('Journal URL') or '').strip(),
                'du': (row.get('URL in DOAJ') or '').strip(),
                'lic': (row.get('Journal license') or '').strip(),
                'apc': (row.get('APC') or '').strip(),
                'fee': (row.get('APC amount') or '').strip(),
                'review': (row.get('Review process') or '').strip(),
                'review_weeks': (row.get('Average number of weeks between article submission and publication') or '').strip(),
            }
            rec = (issn and by_issn.get(issn)) or (eissn and by_issn.get(eissn)) or by_title.get(nt)
            if rec is not None:
                rec['doaj'] = payload
                if not rec.get('publisher') and row.get('Publisher'):
                    rec['publisher'] = row.get('Publisher', '').strip()
                if not rec.get('country') and row.get('Country of publisher'):
                    rec['country'] = row.get('Country of publisher', '').strip()
                matched += 1
            elif store is not None:
                key = 'doaj:' + (issn or eissn or nt)
                rec = store.get(key)
                if rec is None:
                    rec = {
                        'name': title, 'issn': issn, 'eissn': eissn,
                        'publisher': (row.get('Publisher') or '').strip(),
                        'country': (row.get('Country of publisher') or '').strip(),
                        'abbr20': '', 'indices': [], 'wos_categories': [], 'esi_category': '',
                        'doaj_only': True,
                    }
                    store[key] = rec
                    by_title.setdefault(nt, rec)
                    for k in (issn, eissn):
                        if k: by_issn.setdefault(k, rec)
                rec['doaj'] = payload
                standalone += 1
    return matched, standalone


# ───────────────────────── main ─────────────────────────

def main():
    store: dict = {}
    print('== WoS Core ==')
    for idx, fn in INDEX_FILES.items():
        p = LIST_DIR / fn
        if not p.exists():
            print(f'  skip {idx}: missing'); continue
        n = parse_wos_csv(p, idx, store)
        print(f'  {idx}: +{n} (store={len(store)})')

    def rebuild_lookups():
        by_issn, by_title = {}, {}
        for r in store.values():
            for k in (r.get('issn'), r.get('eissn')):
                if k: by_issn.setdefault(k, r)
            by_title.setdefault(norm_title(r['name']), r)
        return by_issn, by_title

    by_issn, by_title = rebuild_lookups()

    print('== JCR 2025 ==')
    hits = parse_jcr(JCR_FILE, store, by_title)
    print(f'  JCR: {hits} (store={len(store)})')
    by_issn, by_title = rebuild_lookups()

    print('== ESI 22 大类 ==')
    hits = parse_esi(ESI_FILE, by_issn, by_title)
    print(f'  ESI matched: {hits}')

    print('== 中科院 2025 完整版 (大类分区) ==')
    h, m = parse_cas(CAS_FILE, by_title)
    print(f'  CAS matched: {h}  unmatched: {m}')

    print('== 长江大学 中文大类 ==')
    h = parse_changjiang(CJU_FILE, by_issn, by_title)
    print(f'  Changjiang matched: {h}')

    print('== 中科院预警名单 2020-2025 ==')
    h = parse_warning_xlsx(LIST_DIR / '国际期刊预警名单_2020-2025.xlsx', by_title, by_issn)
    print(f'  warning matched: {h}')

    print('== ShowJCR JCR 2024 IF ==')
    h = parse_showjcr_if(SHOW_JCR, by_title, by_issn)
    print(f'  IF matched: {h}')

    print('== ShowJCR 中科院分区 2025 (小类分区) ==')
    h = parse_showjcr_fqb(SHOW_FQB, by_title, by_issn)
    print(f'  FQB matched: {h}')

    print('== ShowJCR 新锐版 2026 (中文刊名) ==')
    h = parse_showjcr_xr(SHOW_XR, by_title, by_issn)
    print(f'  XR matched: {h}')

    print('== ShowJCR CCF 2026 ==')
    h = parse_showjcr_ccf(SHOW_CCF, by_title, by_issn)
    print(f'  CCF matched: {h}')

    print('== Scopus Source List ==')
    h, s, inactive = parse_scopus(SCOPUS_FILE, by_title, by_issn, store=store)
    print(f'  Scopus matched: {h}  standalone active: +{s}  inactive matched: {inactive}')
    by_issn, by_title = rebuild_lookups()

    print('== EI Compendex Source List ==')
    h, s, zh = parse_ei_compendex(EI_FILE, by_title, by_issn, store=store)
    print(f'  EI matched: {h}  standalone: +{s}  Chinese-title merged: {zh}')
    by_issn, by_title = rebuild_lookups()

    print('== ABDC Journal Quality List ==')
    abdc_file = first_existing(ABDC_CANDIDATES)
    if abdc_file:
        h = parse_abdc(abdc_file, by_title, by_issn, store=store)
        print(f'  ABDC matched: {h} ({abdc_file.name})')
    else:
        h = 0
        print('  ABDC skipped: file not found')

    print('== ABS / AJG Academic Journal Guide ==')
    abs_file = first_existing(ABS_CANDIDATES)
    if abs_file:
        h = parse_abs(abs_file, by_title, by_issn, store=store)
        print(f'  ABS matched: {h} ({abs_file.name})')
    else:
        print('  ABS skipped: file not found')

    print('== 中国科协 merge ==')
    h = merge_cnkx_to_main(by_issn, by_title)
    print(f'  CNKX merged: {h}')

    print('== OAJ 全球开放获取期刊索引 ==')
    h, s = parse_oaj(OAJ_FILE, by_title, by_issn, store=store)
    print(f'  OAJ matched: {h}  OAJ-only: +{s}')
    by_issn, by_title = rebuild_lookups()

    print('== DOAJ Directory of Open Access Journals ==')
    h, s = parse_doaj(DOAJ_FILE, by_title, by_issn, store=store)
    print(f'  DOAJ matched: {h}  DOAJ-only: +{s}')
    by_issn, by_title = rebuild_lookups()

    # ────── finalize ──────
    for rec in store.values():
        flag = infer_flagship(rec.get('name') or '', rec.get('publisher') or '')
        if flag:
            rec['flagship'] = flag
    journals = list(store.values())
    journals.sort(key=lambda r: r['name'])

    # stats
    idx_c = Counter()
    for r in journals:
        for i in r['indices']: idx_c[i] += 1
    cas_c = Counter(); cas_top = 0
    if_count = 0; warning_count = 0; cn_name_count = 0; ccf_count = 0; abdc_count = 0; abs_count = 0; cnkx_count = 0; scopus_count = 0; ei_count = 0; oaj_count = 0; doaj_count = 0
    for r in journals:
        z = r.get('cas_zone')
        if z: cas_c[z] += 1
        if r.get('cas_top'): cas_top += 1
        if r.get('if_2024'): if_count += 1
        if r.get('warning'): warning_count += 1
        if r.get('cn_name'): cn_name_count += 1
        if r.get('ccf'): ccf_count += 1
        if r.get('abdc'): abdc_count += 1
        if r.get('abs'): abs_count += 1
        if r.get('cnkx'): cnkx_count += 1
        if r.get('scopus') and r.get('scopus', {}).get('active') is not False: scopus_count += 1
        if 'EI' in r.get('indices', []): ei_count += 1
        if r.get('oaj'): oaj_count += 1
        if r.get('doaj'): doaj_count += 1

    print('== Stats ==')
    print(f'  total: {len(journals)}')
    print(f'  indices: {dict(idx_c)}')
    print(f'  CAS zones: {dict(cas_c)} Top={cas_top}')
    print(f'  IF: {if_count}  warning: {warning_count}  中文刊名: {cn_name_count}  CCF: {ccf_count}  ABDC: {abdc_count}  ABS: {abs_count}  CNKX: {cnkx_count}  Scopus: {scopus_count}  EI: {ei_count}  OAJ: {oaj_count}  DOAJ: {doaj_count}')

    # Strip large non-essential fields to stay under CF Pages 25 MB limit
    for r in journals:
        r.pop('address', None)
        r.pop('languages', None)

    # ────── Merge review_cycles (CrossRef) ──────
    RC_FILE = DATA_DIR / 'review_cycles.json'
    if RC_FILE.exists():
        with open(RC_FILE, 'r') as f:
            rc = json.load(f)  # {issn: {median_days, sample_size, source, ...}}
        rc_hits = 0
        for r in journals:
            issn = r.get('issn', '') or ''
            eissn = r.get('eissn', '') or ''
            rcd = rc.get(issn) or rc.get(eissn)
            if rcd:
                r['crossref'] = rcd
                rc_hits += 1
        print(f'  CrossRef review_cycles: {rc_hits} journals matched ({len(rc)} total in file)')
    else:
        print(f'  WARNING: {RC_FILE} not found, skipping CrossRef merge')

    # main write
    with open(DATA_DIR / 'journals.json', 'w', encoding='utf-8') as f:
        json.dump(journals, f, ensure_ascii=False, separators=(',', ':'))

    # gzip for production (CF Pages 25MB per-file limit)
    import gzip
    with open(DATA_DIR / 'journals.json', 'rb') as fin:
        with gzip.open(DATA_DIR / 'journals.json.gz', 'wb', compresslevel=9) as fout:
            shutil.copyfileobj(fin, fout)
    gz_size = (DATA_DIR / 'journals.json.gz').stat().st_size
    print(f'  journals.json.gz: {gz_size/1024/1024:.2f} MB')

    # wos_categories
    wos_c = Counter()
    for r in journals:
        for c in r['wos_categories']: wos_c[c] += 1
    with open(DATA_DIR / 'wos_categories.json', 'w', encoding='utf-8') as f:
        json.dump([{'name': k, 'count': v} for k,v in wos_c.most_common()],
                  f, ensure_ascii=False, indent=2)

    # esi_categories
    esi_c = Counter()
    for r in journals:
        if r['esi_category']: esi_c[r['esi_category']] += 1
    with open(DATA_DIR / 'esi_categories.json', 'w', encoding='utf-8') as f:
        json.dump([{'name': k, 'count': v} for k,v in esi_c.most_common()],
                  f, ensure_ascii=False, indent=2)

    # ────── domestic tab JSON ──────
    domestic = {
        'cnkx': None,
        'zju': None,
        'school_a': None,
        'ccft': [],
        'cssci_core': [],
        'cssci_ext': [],
        'pku_core': [],
        'cnki_major': None,  # full list of CNKI major journals (ISSN + title)
    }
    if CNKX_RECORDS.exists():
        cnkx_records = json.loads(CNKX_RECORDS.read_text(encoding='utf-8'))
        cnkx_by_issn = json.loads(CNKX_JSON.read_text(encoding='utf-8')) if CNKX_JSON.exists() else {}
        cnkx_domains = json.loads(CNKX_DOMAINS.read_text(encoding='utf-8')).get('domains', []) if CNKX_DOMAINS.exists() else []
        domestic['cnkx'] = {'records': cnkx_records, 'by_issn': cnkx_by_issn, 'domains': cnkx_domains}
    if ZJU_JSON.exists():
        domestic['zju'] = json.loads(ZJU_JSON.read_text(encoding='utf-8'))
    if SCHOOL_A_JSON.exists():
        domestic['school_a'] = json.loads(SCHOOL_A_JSON.read_text(encoding='utf-8'))
    if CSSCI_CORE_JSON.exists():
        domestic['cssci_core'] = json.loads(CSSCI_CORE_JSON.read_text(encoding='utf-8'))
    if CSSCI_EXT_JSON.exists():
        domestic['cssci_ext'] = json.loads(CSSCI_EXT_JSON.read_text(encoding='utf-8'))
    if PKU_CORE_JSON.exists():
        domestic['pku_core'] = json.loads(PKU_CORE_JSON.read_text(encoding='utf-8'))
    if SHOW_CCFT.exists():
        with open(SHOW_CCFT, 'r', encoding='utf-8-sig', newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                domestic['ccft'].append({
                    'cn_name': (row.get('中文刊名') or '').strip(),
                    'en_name': (row.get('Journal') or '').strip(),
                    'cn_code': (row.get('CN号') or '').strip(),
                    'language': (row.get('语种') or '').strip(),
                    'org': (row.get('主办单位') or '').strip(),
                    'ccf_area': (row.get('CCF推荐类别') or '').strip(),
                    'tier': (row.get('T分区') or '').strip(),
                })
    # ────── CNKI Major Journals (全量中文期刊主目录) ──────
    cnki_major_records = []
    cnki_major_by_issn = {}
    if CNKI_MAJOR_FILE.exists():
        with open(CNKI_MAJOR_FILE, 'r', encoding='utf-8-sig', newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                title = (row.get('title') or '').strip()
                if not title: continue
                issn = clean_issn(row.get('issn') or '')
                cn = (row.get('cn') or '').strip()
                sponsor = (row.get('sponsor') or '').strip()
                compound_if = (row.get('compound_if') or '').strip()
                comprehensive_if = (row.get('comprehensive_if') or '').strip()
                tags = (row.get('tags') or '').strip()
                categories = (row.get('major_categories') or '').strip()
                rec = {
                    'name': title,
                    'issn': issn,
                    'cn_code': cn,
                    'sponsor': sponsor,
                    'compound_if': compound_if,
                    'comprehensive_if': comprehensive_if,
                    'tags': tags,
                    'major_categories': [c.strip() for c in categories.split('|') if c.strip()],
                }
                cnki_major_records.append(rec)
                if issn:
                    cnki_major_by_issn.setdefault(issn, []).append(rec)
        domestic['cnki_major'] = {'records': cnki_major_records, 'by_issn': cnki_major_by_issn}
        print(f'  CNKI Major: {len(cnki_major_records)} records, {len(cnki_major_by_issn)} with ISSN')
        # Also tag international journals that appear in CNKI Major
        cnki_tagged = 0
        for r in journals:
            for issn_key in (r.get('issn',''), r.get('eissn','')):
                if issn_key and issn_key in cnki_major_by_issn:
                    r['cnki_major'] = True
                    cnki_tagged += 1
                    break
        print(f'  CNKI Major tagged on international journals: {cnki_tagged}')
    with open(DATA_DIR / 'domestic.json', 'w', encoding='utf-8') as f:
        json.dump(domestic, f, ensure_ascii=False)

    # meta
    meta = {
        'source': 'WoS Core + JCR 2025 + ESI + 中科院 2025 + 长江大学 + ShowJCR (JCR/FQB/XR/CCF/Warning) + Scopus (auto-updated) + EI Compendex Oct. 2025 + ABDC optional + ABS AJG + 中国科协 + OAJ 2025 + DOAJ Journal CSV',
        'last_updated_source': 'WoS Core 2026-04-20',
        'total': len(journals),
        'indices': dict(idx_c),
        'with_if_2024': if_count,
        'with_cas_zone': sum(cas_c.values()),
        'with_cas_top': cas_top,
        'with_warning': warning_count,
        'with_cn_name': cn_name_count,
        'with_ccf': ccf_count,
        'with_abdc': abdc_count,
        'with_abs': abs_count,
        'with_oaj': oaj_count,
        'with_doaj': doaj_count,
        'with_cnkx': cnkx_count,
        'with_scopus': scopus_count,
        'with_ei': ei_count,
        'wos_categories': len(wos_c),
        'esi_categories': len(esi_c),
    }
    with open(DATA_DIR / 'meta.json', 'w', encoding='utf-8') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

    size = (DATA_DIR / 'journals.json').stat().st_size
    print(f'  journals.json: {size/1024/1024:.2f} MB')

    print('== OpenAlex Merge (oa.json) ==')
    merge_script = ROOT / 'scripts' / 'merge_openalex.py'
    if merge_script.exists() and (DATA_DIR / 'openalex_cache.json').exists():
        r = subprocess.run(
            [sys.executable, str(merge_script)],
            cwd=ROOT, capture_output=True, text=True, timeout=120)
        print(r.stdout)
        if r.returncode != 0:
            print(f'  WARN: merge_openalex.py exited {r.returncode}: {r.stderr.strip()}')
        # gzip oa.json for CF Pages 25MB per-file limit
        oa_path = DATA_DIR / 'oa.json'
        if oa_path.exists():
            oa_size = oa_path.stat().st_size
            with open(oa_path, 'rb') as fin:
                with gzip.open(DATA_DIR / 'oa.json.gz', 'wb', compresslevel=9) as fout:
                    shutil.copyfileobj(fin, fout)
            gz_size = (DATA_DIR / 'oa.json.gz').stat().st_size
            print(f'  oa.json: {oa_size/1024/1024:.1f} MB → oa.json.gz: {gz_size/1024/1024:.1f} MB')
    else:
        print('  skip: merge_openalex.py or openalex_cache.json not found')
    return 0


if __name__ == '__main__':
    sys.exit(main())
