#!/usr/bin/env python3
"""Synchronize EI/Compendex coverage without rebuilding unrelated enrichments.

The full journal build has several post-processing stages.  This focused updater
keeps those enriched fields intact while replacing EI membership with the
current official Compendex source list.
"""
from __future__ import annotations

import gzip
import json
import re
import unicodedata
from collections import Counter
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
DEFAULT_SOURCE = ROOT / "list" / "CPXSourceList_072026.xlsx"
SOURCE_LABEL = "Compendex Source List Jul. 9, 2026"


def clean_issn(value) -> str:
    compact = re.sub(r"[^0-9X]", "", str(value or "").upper())
    return f"{compact[:4]}-{compact[4:]}" if len(compact) == 8 else ""


def norm_title(value) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch)).lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def slugify(value, issn, used) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch)).lower()
    base = re.sub(r"[^a-z0-9]+", "-", text).strip("-")[:72].rstrip("-")
    base = base or re.sub(r"[^0-9X]", "", str(issn or "").upper()) or "journal"
    candidate = base
    suffix = re.sub(r"[^0-9X]", "", str(issn or "").upper())
    if candidate in used and suffix:
        candidate = f"{base[:60].rstrip('-')}-{suffix}"
    serial = 2
    root = candidate
    while candidate in used:
        candidate = f"{root}-{serial}"
        serial += 1
    used.add(candidate)
    return candidate


def load_source(path: Path) -> list[dict]:
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    chinese = {}
    if "CHINESE JRS on SERIALS LIST" in book.sheetnames:
        for row in book["CHINESE JRS on SERIALS LIST"].iter_rows(min_row=3, values_only=True):
            if not row:
                continue
            issn, eissn = clean_issn(row[0]), clean_issn(row[1])
            payload = {
                "cn_name": str(row[2] or "").strip(),
                "status": str(row[6] or "").strip(),
            }
            for key in (issn, eissn):
                if key:
                    chinese[key] = payload

    records = []
    for row in book["SERIALS"].iter_rows(min_row=3, values_only=True):
        if not row or str(row[1] or "").strip().lower() != "journal":
            continue
        issn, eissn = clean_issn(row[2]), clean_issn(row[3])
        cn = chinese.get(issn) or chinese.get(eissn) or {}
        subjects = [str(value).strip() for value in row[7:15]
                    if value and str(value).strip() not in {"-", "—"}]
        records.append({
            "title": str(row[0] or "").strip(),
            "issn": issn,
            "eissn": eissn,
            "publisher": str(row[4] or "").strip(),
            "country": str(row[5] or "").strip(),
            "language": str(row[6] or "").strip(),
            "subjects": subjects,
            "cn_name": cn.get("cn_name", ""),
            "status": cn.get("status") or SOURCE_LABEL,
        })
    return records


def load_discontinued_source(path: Path) -> list[dict]:
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "DISCONTINUED" not in book.sheetnames:
        return []
    records = []
    for row in book["DISCONTINUED"].iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue
        records.append({
            "title": str(row[0] or "").strip(),
            "issn": clean_issn(row[1] if len(row) > 1 else ""),
            "eissn": clean_issn(row[2] if len(row) > 2 else ""),
            "publisher": str(row[3] or "").strip() if len(row) > 3 else "",
            "final_year": str(row[4] or "").strip() if len(row) > 4 else "",
            "final_volume": str(row[5] or "").strip() if len(row) > 5 else "",
            "final_issue": str(row[6] or "").strip() if len(row) > 6 else "",
            "final_pages": str(row[7] or "").strip() if len(row) > 7 else "",
        })
    return records


def read_json(path: Path):
    if path.suffix == ".gz":
        with gzip.open(path, "rt", encoding="utf-8") as handle:
            return json.load(handle)
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload) -> None:
    raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    if path.suffix == ".gz":
        with path.open("wb") as target:
            with gzip.GzipFile(fileobj=target, mode="wb", compresslevel=9, mtime=0) as handle:
                handle.write(raw)
    else:
        path.write_bytes(raw)


def sync_records(records: list[dict], source_rows: list[dict], historical_rows: list[dict]) -> tuple[list[dict], dict]:
    old_ei = sum("EI" in (rec.get("indices") or []) for rec in records)
    used_slugs = {str(rec.get("slug") or "") for rec in records if rec.get("slug")}
    by_issn, by_title = {}, {}

    for rec in records:
        rec["indices"] = [item for item in (rec.get("indices") or []) if item != "EI"]
        for key in ("ei_subjects", "ei_status", "ei_historical"):
            rec.pop(key, None)
        rec.pop("ei_historical_only", None)
        for key in (clean_issn(rec.get("issn")), clean_issn(rec.get("eissn"))):
            if key:
                by_issn.setdefault(key, rec)
        title_key = norm_title(rec.get("name") or rec.get("cn_name"))
        if title_key:
            by_title.setdefault(title_key, rec)

    matched = added = 0
    for row in source_rows:
        rec = next((by_issn[key] for key in (row["issn"], row["eissn"])
                    if key and key in by_issn), None)
        if rec is None:
            rec = by_title.get(norm_title(row["title"]))
        if rec is None:
            rec = {
                "name": row["title"],
                "issn": row["issn"],
                "eissn": row["eissn"],
                "publisher": row["publisher"],
                "country": row["country"],
                "abbr20": "",
                "indices": [],
                "wos_categories": [],
                "esi_category": "",
                "ei_only": True,
            }
            rec["slug"] = slugify(row["title"], row["issn"] or row["eissn"], used_slugs)
            records.append(rec)
            added += 1
        else:
            matched += 1
        if "EI" not in rec["indices"]:
            rec["indices"].append("EI")
        rec["ei_subjects"] = row["subjects"]
        rec["ei_status"] = row["status"]
        if row["cn_name"]:
            rec["cn_name"] = row["cn_name"]
        for field in ("issn", "eissn", "publisher", "country"):
            if not rec.get(field) and row.get(field):
                rec[field] = row[field]
        if row["language"] and not rec.get("languages"):
            rec["languages"] = row["language"]
        for key in (row["issn"], row["eissn"]):
            if key:
                by_issn.setdefault(key, rec)
        by_title.setdefault(norm_title(row["title"]), rec)

    historical_matched = historical_added = 0
    for row in historical_rows:
        rec = next((by_issn[key] for key in (row["issn"], row["eissn"])
                    if key and key in by_issn), None)
        if rec is None:
            rec = by_title.get(norm_title(row["title"]))
        if rec is not None and "EI" in (rec.get("indices") or []):
            continue
        if rec is None:
            rec = {
                "name": row["title"],
                "issn": row["issn"],
                "eissn": row["eissn"],
                "publisher": row["publisher"],
                "country": "",
                "abbr20": "",
                "indices": [],
                "wos_categories": [],
                "esi_category": "",
                "ei_historical_only": True,
            }
            rec["slug"] = slugify(row["title"], row["issn"] or row["eissn"], used_slugs)
            records.append(rec)
            historical_added += 1
        else:
            historical_matched += 1
        year = row["final_year"]
        rec["ei_historical"] = {
            "status": "discontinued",
            "final_year": year,
            "final_volume": row["final_volume"],
            "final_issue": row["final_issue"],
            "final_pages": row["final_pages"],
            "source": SOURCE_LABEL,
        }
        rec["ei_status"] = f"Discontinued{f' (final coverage {year})' if year else ''}"
        if row["publisher"] and not rec.get("publisher"):
            rec["publisher"] = row["publisher"]
        for key in (row["issn"], row["eissn"]):
            if key:
                by_issn.setdefault(key, rec)
        by_title.setdefault(norm_title(row["title"]), rec)

    retained = []
    removed = 0
    for rec in records:
        if rec.get("ei_only") and "EI" not in (rec.get("indices") or []):
            if rec.get("ei_historical"):
                rec.pop("ei_only", None)
                rec["ei_historical_only"] = True
                retained.append(rec)
                continue
            has_other_source = any(rec.get(key) for key in (
                "scopus", "doaj", "oaj", "fsta", "cabi", "inspec", "specialty_only",
            ))
            if not has_other_source:
                removed += 1
                continue
            rec.pop("ei_only", None)
        retained.append(rec)
    current_ei = sum("EI" in (rec.get("indices") or []) for rec in retained)
    return retained, {
        "before": old_ei,
        "after": current_ei,
        "matched_source_rows": matched,
        "added": added,
        "removed_obsolete_ei_only": removed,
        "historical_source_rows": len(historical_rows),
        "historical_matched": historical_matched,
        "historical_added": historical_added,
    }


def update_index(records: list[dict]) -> int:
    path = DATA / "journal_index.json"
    index = read_json(path)
    added = 0
    for rec in records:
        if not (rec.get("ei_only") or rec.get("ei_historical_only")):
            continue
        slug = str(rec.get("slug") or "")
        if not slug:
            continue
        if slug not in index:
            entry = {
                "n": rec.get("name") or "",
                "i": rec.get("issn") or "",
                "is": rec.get("eissn") or "",
                "p": rec.get("publisher") or "",
                "ix": ["EI"] if rec.get("ei_only") else [],
            }
            index[slug] = {key: value for key, value in entry.items() if value not in ("", [])}
            added += 1
        for value in (rec.get("issn"), rec.get("eissn")):
            compact = re.sub(r"[^0-9X]", "", str(value or "").upper())
            if compact and compact not in index:
                index[compact] = {"_r": slug}
    write_json(path, index)
    return added


def main() -> None:
    source_rows = load_source(DEFAULT_SOURCE)
    historical_rows = load_discontinued_source(DEFAULT_SOURCE)
    results = {}
    full_records = None
    for filename in ("journals.json.gz", "journals_light.json.gz"):
        path = DATA / filename
        records, stats = sync_records(read_json(path), source_rows, historical_rows)
        if filename == "journals_light.json.gz" and full_records is not None:
            full_slugs = {str(rec.get("slug") or "") for rec in full_records}
            before_align = len(records)
            records = [rec for rec in records if str(rec.get("slug") or "") in full_slugs]
            stats["removed_not_in_full"] = before_align - len(records)
        write_json(path, records)
        results[filename] = {"records": len(records), **stats}
        if filename == "journals.json.gz":
            full_records = records

    plain_path = DATA / "journals.json"
    if plain_path.exists():
        plain_records, plain_stats = sync_records(read_json(plain_path), source_rows, historical_rows)
        write_json(plain_path, plain_records)
        results[plain_path.name] = {"records": len(plain_records), **plain_stats}

    meta_path = DATA / "meta.json"
    meta = read_json(meta_path)
    current_ei = results["journals.json.gz"]["after"]
    meta["total"] = len(full_records or [])
    meta.setdefault("indices", {})["EI"] = current_ei
    meta["with_ei"] = current_ei
    meta["ei_source_updated"] = "2026-07-09"
    meta["ei_source_url"] = "https://www.elsevier.com/products/engineering-village/databases/compendex"
    meta["source"] = re.sub(
        r"EI Compendex (?:Oct\. 2025|Jul\. 9, 2026)",
        "EI Compendex Jul. 9, 2026",
        str(meta.get("source") or ""),
    )
    meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    results["source_rows"] = len(source_rows)
    results["journal_index_added"] = update_index(full_records or [])
    results["ei_statuses"] = dict(Counter(
        rec.get("ei_status") for rec in (full_records or []) if rec.get("ei_status")
    ))
    print(json.dumps(results, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
