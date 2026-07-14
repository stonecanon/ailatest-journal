#!/usr/bin/env python3
"""Merge public specialty-index title lists into deployed journal datasets.

Supported inputs:
- EBSCO FSTA with Full Text coverage HTML (public title list)
- IET Inspec active source list XLSX (public source list)
- CABI CAB Abstracts serial cited report XLS (official source report)

The FSTA list is intentionally stored as ``fsta_full_text``.  It is a subset
of FSTA, not the complete FSTA abstracting and indexing database.
"""
from __future__ import annotations

import argparse
import gzip
import json
import re
import unicodedata
import urllib.request
from html.parser import HTMLParser
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

try:
    import xlrd
except ImportError:  # pragma: no cover
    xlrd = None


ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
FSTA_URL = "https://about.ebsco.com/m/ee/Marketing/titleLists/fwt-coverage.htm"
CABI_URL = "https://www.cabi.org/what-we-do/publishing-products/"


def clean_issn(value) -> str:
    raw = re.sub(r"[^0-9X]", "", str(value or "").upper())
    return f"{raw[:4]}-{raw[4:]}" if len(raw) == 8 else ""


class TableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.in_table = self.in_cell = False
        self.rows: list[list[str]] = []
        self.row: list[str] | None = None
        self.cell: list[str] = []

    def handle_starttag(self, tag, attrs):
        attrs = dict(attrs)
        if tag == "table" and attrs.get("id") == "dataTable":
            self.in_table = True
        elif self.in_table and tag == "tr":
            self.row = []
        elif self.in_table and tag in {"td", "th"} and self.row is not None:
            self.in_cell = True
            self.cell = []

    def handle_data(self, data):
        if self.in_cell:
            self.cell.append(data)

    def handle_endtag(self, tag):
        if self.in_table and tag in {"td", "th"} and self.in_cell:
            self.row.append(" ".join("".join(self.cell).split()))
            self.in_cell = False
        elif self.in_table and tag == "tr" and self.row is not None:
            if self.row:
                self.rows.append(self.row)
            self.row = None
        elif tag == "table" and self.in_table:
            self.in_table = False


def load_fsta(source: str) -> list[dict]:
    if re.match(r"^https?://", source):
        req = urllib.request.Request(source, headers={"User-Agent": "AILatest-Journal/1.0"})
        html = urllib.request.urlopen(req, timeout=45).read().decode("utf-8", "replace")
    else:
        html = Path(source).read_text(encoding="utf-8", errors="replace")
    parser = TableParser()
    parser.feed(html)
    if not parser.rows:
        raise RuntimeError("FSTA coverage table not found")
    header, *rows = parser.rows
    out = []
    for row in rows:
        row += [""] * (len(header) - len(row))
        rec = dict(zip(header, row))
        issn = clean_issn(rec.get("ISSN"))
        if not issn:
            continue
        out.append({
            "issn": issn,
            "title": rec.get("Publication Name", ""),
            "publisher": rec.get("Publisher", ""),
            "country": rec.get("Country", ""),
            "source_type": rec.get("Source Type", ""),
            "peer_reviewed": rec.get("Peer-Reviewed") == "Y",
            "full_text_start": rec.get("Full Text Start", ""),
            "full_text_stop": rec.get("Full Text Stop", ""),
            "availability": rec.get("Availability*", ""),
        })
    return out


def find_header_row(sheet):
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        vals = [str(v or "").strip() for v in row]
        joined = "|".join(vals).lower()
        if "journal title" in joined and ("pissn" in joined or "issn" in joined):
            return idx, vals
    raise RuntimeError("Inspec header row not found")


def load_inspec(path: str) -> list[dict]:
    if not openpyxl:
        raise RuntimeError("openpyxl is required for the Inspec XLSX")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    header_row, header = find_header_row(sheet)
    keys = {re.sub(r"\s+", " ", h.lower()).strip(): i for i, h in enumerate(header)}

    def value(row, *names):
        for name in names:
            idx = keys.get(name)
            if idx is not None and idx < len(row):
                return str(row[idx] or "").strip()
        return ""

    out, seen = [], set()
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        title = value(row, "journal title", "source title")
        pissn = clean_issn(value(row, "pissn", "print issn", "issn"))
        eissn = clean_issn(value(row, "eissn", "electronic issn"))
        if not title or not (pissn or eissn):
            continue
        key = (pissn, eissn, title.casefold())
        if key in seen:
            continue
        seen.add(key)
        out.append({
            "title": title,
            "issn": pissn,
            "eissn": eissn,
            "publisher": value(row, "publisher"),
            "country": value(row, "publisher country", "country"),
            "inspec_id": value(row, "inspec id"),
        })
    return out


def load_cabi(path: str) -> list[dict]:
    """Read CABI's CAB Abstracts serial cited report (legacy BIFF XLS)."""
    if not xlrd:
        raise RuntimeError("xlrd is required for the CABI XLS")
    book = xlrd.open_workbook(path, on_demand=True)
    sheet = book.sheet_by_index(0)
    header = [str(v or "").strip() for v in sheet.row_values(0)]
    out, seen = [], set()
    for row_no in range(1, sheet.nrows):
        values = sheet.row_values(row_no)
        rec = dict(zip(header, values))
        title = str(rec.get("Document Title") or "").strip()
        issn = clean_issn(rec.get("ISSN"))
        eissn = clean_issn(rec.get("eISSN"))
        if not title or not (issn or eissn):
            continue
        key = (issn, eissn, title.casefold())
        if key in seen:
            continue
        seen.add(key)
        out.append({
            "title": title,
            "issn": issn,
            "eissn": eissn,
            "publisher": str(rec.get("Publisher name") or "").strip(),
            "country": str(rec.get("Country of publication") or "").strip(),
        })
    book.release_resources()
    return out


def issn_keys(rec: dict) -> set[str]:
    return {x for x in (clean_issn(rec.get("issn")), clean_issn(rec.get("eissn"))) if x}


def norm_title(value: str) -> str:
    value = unicodedata.normalize("NFKD", str(value or "")).casefold()
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", " ", value).strip()


def slugify(title: str, issn: str, used: set[str]) -> str:
    value = unicodedata.normalize("NFKD", title).encode("ascii", "ignore").decode().lower()
    base = re.sub(r"[^a-z0-9]+", "-", value).strip("-")[:60].rstrip("-")
    base = base or re.sub(r"[^0-9X]", "", issn.upper()) or "journal"
    slug = base
    if slug in used:
        suffix = re.sub(r"[^0-9X]", "", issn.upper()) or str(len(used) + 1)
        slug = f"{base[:50].rstrip('-')}-{suffix}"
    serial = 2
    candidate = slug
    while candidate in used:
        candidate = f"{slug}-{serial}"
        serial += 1
    used.add(candidate)
    return candidate


def merge_source(records: list[dict], rows: list[dict], flag: str, aliases=()) -> dict:
    """Match by ISSN/EISSN, then exact normalized title; append unmatched serials."""
    flags = (flag, *aliases)
    for rec in records:
        for key in flags:
            rec.pop(key, None)

    by_issn, by_title = {}, {}
    used_slugs = {str(r.get("slug") or "") for r in records if r.get("slug")}
    for rec in records:
        for key in issn_keys(rec):
            by_issn.setdefault(key, rec)
        title_key = norm_title(rec.get("name") or rec.get("en_name") or rec.get("cn_name"))
        if title_key:
            by_title.setdefault(title_key, rec)

    matched = added = 0
    for row in rows:
        keys = issn_keys(row)
        rec = next((by_issn[k] for k in keys if k in by_issn), None)
        if rec is None:
            rec = by_title.get(norm_title(row.get("title")))
        if rec is None:
            title = str(row.get("title") or "").strip()
            issn = clean_issn(row.get("issn"))
            eissn = clean_issn(row.get("eissn"))
            if not title or not (issn or eissn):
                continue
            rec = {
                "name": title,
                "issn": issn,
                "eissn": eissn,
                "publisher": str(row.get("publisher") or "").strip(),
                "country": str(row.get("country") or "").strip(),
                "abbr20": "",
                "indices": [],
                "wos_categories": [],
                "esi_category": "",
                "specialty_only": True,
            }
            rec["slug"] = slugify(title, issn or eissn, used_slugs)
            records.append(rec)
            added += 1
        else:
            matched += 1
            for key in ("issn", "eissn", "publisher", "country"):
                if not rec.get(key) and row.get(key):
                    rec[key] = row[key]
        for key in flags:
            rec[key] = True
        for key in issn_keys(rec):
            by_issn.setdefault(key, rec)
        by_title.setdefault(norm_title(rec.get("name")), rec)
    return {"matched": matched, "added": added, "total": matched + added}


def read_json(path: Path):
    if path.suffix == ".gz":
        with gzip.open(path, "rt", encoding="utf-8") as f:
            return json.load(f)
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, data):
    raw = json.dumps(data, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    if path.suffix == ".gz":
        with gzip.open(path, "wb", compresslevel=9) as f:
            f.write(raw)
    else:
        path.write_bytes(raw)


def update_routing(records: list[dict]) -> dict:
    """Make newly added specialty journals resolvable by detail URL and sitemap."""
    index_path = DATA / "journal_index.json"
    index = read_json(index_path) if index_path.exists() else {}
    added_index = 0
    sitemap_urls = []
    for rec in records:
        if not rec.get("specialty_only"):
            continue
        slug = str(rec.get("slug") or "").strip()
        if not slug:
            continue
        if slug not in index:
            entry = {
                "n": rec.get("name") or "",
                "i": rec.get("issn") or "",
                "is": rec.get("eissn") or "",
                "p": rec.get("publisher") or "",
                "ix": rec.get("indices") or [],
                "sp": [name for name, enabled in (
                    ("FSTA", rec.get("fsta")),
                    ("CABI", rec.get("cabi")),
                    ("Inspec", rec.get("inspec")),
                ) if enabled],
            }
            index[slug] = {k: v for k, v in entry.items() if v not in ("", [], None)}
            added_index += 1
        for value in (rec.get("issn"), rec.get("eissn")):
            compact = re.sub(r"[^0-9X]", "", str(value or "").upper())
            if compact and compact not in index:
                index[compact] = {"_r": slug}
        sitemap_urls.append(f"https://journal.ailatest.org/journal/{slug}/")
    write_json(index_path, index)

    sitemap_path = ROOT / "sitemap.xml"
    added_sitemap = 0
    if sitemap_path.exists():
        xml = sitemap_path.read_text(encoding="utf-8")
        additions = []
        for url in sitemap_urls:
            if f"<loc>{url}</loc>" not in xml:
                additions.append(f"  <url><loc>{url}</loc><priority>0.8</priority></url>")
        if additions:
            xml = xml.replace("</urlset>", "\n".join(additions) + "\n</urlset>")
            sitemap_path.write_text(xml, encoding="utf-8")
            added_sitemap = len(additions)
    return {"journal_index_added": added_index, "sitemap_added": added_sitemap}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--fsta", default=FSTA_URL, help="FSTA coverage HTML URL or file")
    ap.add_argument("--inspec", help="IET Inspec active source list XLSX")
    ap.add_argument("--cabi", help="CABI CAB Abstracts serial cited report XLS")
    args = ap.parse_args()

    fsta_rows = load_fsta(args.fsta)
    inspec_rows = load_inspec(args.inspec) if args.inspec else []
    cabi_rows = load_cabi(args.cabi) if args.cabi else []
    active_fsta = [r for r in fsta_rows if not r.get("full_text_stop")]

    stats = {}
    full_records = None
    # The live site consumes the compressed full/light datasets.  The legacy
    # uncompressed deploy snapshot is intentionally left untouched because it
    # is no longer served and would exceed the host's per-file size limit.
    for name in ("journals.json.gz", "journals_light.json.gz"):
        path = DATA / name
        if not path.exists():
            continue
        records = read_json(path)
        source_stats = {
            "fsta": merge_source(records, active_fsta, "fsta", ("fsta_full_text",)),
            "inspec": merge_source(records, inspec_rows, "inspec") if inspec_rows else {},
            "cabi": merge_source(records, cabi_rows, "cabi") if cabi_rows else {},
        }
        write_json(path, records)
        stats[name] = {"records": len(records), **source_stats}
        if name == "journals.json.gz":
            full_records = records

    routing = update_routing(full_records or [])

    payload = {
        "updated": "2026-07-14",
        "sources": {
            "fsta_full_text": {
                "label": "FSTA with Full Text",
                "scope": "full_text_subset",
                "url": FSTA_URL,
                "records": len(fsta_rows),
                "active_records": len(active_fsta),
                "active_issns": len({x for r in active_fsta for x in issn_keys(r)}),
            },
            "inspec": {
                "label": "Inspec",
                "scope": "active_source_list",
                "url": "https://www.theiet.org/publishing/solutions-for-your-institution-or-organisation/inspec/inspec-content-and-coverage",
                "records": len(inspec_rows),
                "issns": len({x for r in inspec_rows for x in issn_keys(r)}),
            },
            "cabi": {
                "label": "CAB Abstracts",
                "scope": "serial_cited_report",
                "report_date": "2013-09",
                "url": CABI_URL,
                "records": len(cabi_rows),
                "issns": len({x for r in cabi_rows for x in issn_keys(r)}),
            },
        },
        "matches": stats,
        "routing": routing,
    }
    write_json(DATA / "specialty_indexes.json", payload)
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
