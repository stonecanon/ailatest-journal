#!/usr/bin/env python3
"""Update journal impact factors/JCR quartiles from the latest JCR workbook.

The public UI still has many references to the historical ``if_2024`` field.
For backward compatibility this script writes the latest JIF into ``if_2024``
as well as explicit year-aware fields such as ``if_2025`` and ``if_history``.
"""
from __future__ import annotations

import argparse
import gzip
import json
import re
import shutil
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required to read the JCR workbook") from exc


ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
DEFAULT_WORKBOOKS = [
    ROOT / "list" / "2026影响因子和分区_所有期刊.xlsx",
    ROOT / "2026影响因子和分区_所有期刊.xlsx",
    Path("/Users/zhizhi/百度同步/AI 工作区/00_每日更新/ailatest-journal/list/2026影响因子和分区_所有期刊.xlsx"),
]
JOURNALS_GZ = DATA_DIR / "journals.json.gz"
JOURNALS_JSON = DATA_DIR / "journals.json"
LIGHT_GZ = DATA_DIR / "journals_light.json.gz"
LIGHT_V2_GZ = DATA_DIR / "journals_light_v2.json.gz"
ANNUAL_OUTPUT_CANDIDATES = [
    DATA_DIR / "annual_outputs.json.gz",
    DATA_DIR / "annual_outputs (1).json.gz",
]
AUTHOR_FREE_OA_LABELS = {"diamond", "hybrid", "subscription_paid_read"}


def norm_title(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value)).upper()
    return re.sub(r"[^A-Z0-9]+", "", text)


def clean_issn(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().upper()
    match = re.search(r"\b(\d{4})-?(\d{3}[\dX])\b", text)
    return f"{match.group(1)}-{match.group(2)}" if match else ""


def number_or_none(value: object):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def int_or_none(value: object):
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def read_json_gz(path: Path):
    return json.loads(gzip.decompress(path.read_bytes()).decode("utf-8"))


def write_json_bundle(path_json: Path, path_gz: Path, data) -> None:
    path_json.write_text(json.dumps(data, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    with path_json.open("rb") as fin, gzip.open(path_gz, "wb", compresslevel=9) as fout:
        shutil.copyfileobj(fin, fout)


def write_json_gz(path_gz: Path, data) -> None:
    with gzip.open(path_gz, "wt", encoding="utf-8", compresslevel=9) as fout:
        json.dump(data, fout, ensure_ascii=False, separators=(",", ":"))


def first_existing(paths: list[Path]) -> Path | None:
    for path in paths:
        if path.exists():
            return path
    return None


def build_lookup(journals: list[dict]):
    by_issn: dict[str, dict] = {}
    by_title: dict[str, dict] = {}
    for rec in journals:
        for key in ("issn", "eissn"):
            issn = clean_issn(rec.get(key))
            if issn:
                by_issn.setdefault(issn, rec)
        for key in ("name", "en_name", "cn_name", "abbr20"):
            nt = norm_title(rec.get(key))
            if nt:
                by_title.setdefault(nt, rec)
    return by_issn, by_title


def read_latest_jcr(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Journals" not in wb.sheetnames:
        raise SystemExit(f"{path} does not contain a Journals sheet")
    ws = wb["Journals"]
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    idx = {name: i for i, name in enumerate(headers)}

    def col(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    out = []
    for row in rows:
        title = (col(row, "Journal name") or "").strip() if isinstance(col(row, "Journal name"), str) else col(row, "Journal name")
        if not title:
            continue
        jif = number_or_none(col(row, "2025 JIF"))
        # Some JCR workbooks expose the latest JIF without self-cites as a
        # separate column.  Keep the parser tolerant because Clarivate has
        # changed the header spelling between exports.
        jif_without_self = number_or_none(
            next((col(row, name) for name in (
                "2025 JIF without self-cites",
                "2025 JIF Without Self-Cites",
                "2025 JIF without self-citations",
                "2025 JIF Without Self-Citations",
                "2025 JIF without self citations",
                "JIF without self-cites",
                "JIF Without Self-Cites",
                "JIF without self-citations",
                "JIF Without Self-Citations",
            ) if col(row, name) not in (None, "")), None)
        )
        q = str(col(row, "JIF quartile") or "").strip().upper()
        rank = str(col(row, "JIF rank") or "").strip()
        cats = str(col(row, "Categories") or "").strip()
        cat_list = [c.strip() for c in re.split(r"[;；]", cats) if c and c.strip()]
        rec = {
            "title": str(title).strip(),
            "issn": clean_issn(col(row, "ISSN")),
            "eissn": clean_issn(col(row, "eISSN")),
            "jif": jif,
            "jif_without_self": jif_without_self,
            "quartile": q if q in {"Q1", "Q2", "Q3", "Q4"} else "",
            "rank": rank,
            "categories": cat_list,
            "jcr_year": int(col(row, "JCR year") or 2025),
            "five_year_jif": number_or_none(col(row, "5-year JIF")),
            "jci": number_or_none(col(row, "JCI")),
            "total_articles": int_or_none(col(row, "Total articles")),
            "total_citations": int_or_none(col(row, "Total citations")),
        }
        out.append(rec)
    return out


def output_history_for(rec: dict, annual_outputs: dict, years: int = 12) -> list[dict]:
    for key in (clean_issn(rec.get("issn")), clean_issn(rec.get("eissn"))):
        if key and key in annual_outputs:
            raw = annual_outputs[key]
            items = []
            for year, count in raw.items():
                try:
                    y = int(year)
                    c = int(count)
                except (TypeError, ValueError):
                    continue
                items.append({"year": y, "count": c})
            items.sort(key=lambda x: x["year"])
            return items[-years:]
    return []


def attach_free_flags(journals: list[dict]) -> int:
    oa_path = DATA_DIR / "oa.json.gz"
    if not oa_path.exists():
        return 0
    oa_data = read_json_gz(oa_path)
    count = 0
    for rec in journals:
        rec.pop("free", None)
        for key in (clean_issn(rec.get("issn")), clean_issn(rec.get("eissn"))):
            if key and str(oa_data.get(key, {}).get("l", "")).lower() in AUTHOR_FREE_OA_LABELS:
                rec["free"] = True
                count += 1
                break
    return count


def apply_updates(journals: list[dict], latest_rows: list[dict], annual_outputs: dict) -> tuple[int, int]:
    by_issn, by_title = build_lookup(journals)
    matched = 0
    for row in latest_rows:
        rec = (
            (row["issn"] and by_issn.get(row["issn"]))
            or (row["eissn"] and by_issn.get(row["eissn"]))
            or by_title.get(norm_title(row["title"]))
        )
        if rec is None:
            continue

        old_if = number_or_none(rec.get("if_2024"))
        history = {}
        if isinstance(rec.get("if_history"), dict):
            for y, v in rec["if_history"].items():
                nv = number_or_none(v)
                if nv is not None:
                    history[str(y)] = nv
        if old_if is not None and "2024" not in history:
            history["2024"] = old_if
        if row["jif"] is not None:
            history["2025"] = row["jif"]
            rec["if_2024"] = row["jif"]
            rec["if_2025"] = row["jif"]
            rec["if_latest"] = row["jif"]
            rec["if_latest_year"] = 2025
        if row.get("jif_without_self") is not None:
            rec["jif_without_self_cites_2025"] = row["jif_without_self"]
            if row.get("jif") not in (None, 0):
                rec["self_citation_rate_2025"] = round(
                    min(100.0, max(0.0, (row["jif"] - row["jif_without_self"]) / row["jif"] * 100)),
                    2,
                )

        if history:
            rec["if_history"] = dict(sorted(history.items(), key=lambda item: int(item[0])))
        if row["quartile"]:
            rec["if_quartile"] = row["quartile"]
        if row["rank"]:
            rec["if_rank"] = row["rank"]
        if row["categories"]:
            rec["jcr_cat"] = row["categories"][0]
            rec["jcr_cats"] = row["categories"]
        rec["jcr_year"] = row["jcr_year"]
        rec["jcr_release_year"] = 2026
        if row["five_year_jif"] is not None:
            rec["five_year_if_2025"] = row["five_year_jif"]
        if row["jci"] is not None:
            rec["jci_2025"] = row["jci"]
        if row["total_articles"] is not None:
            rec["jcr_total_articles_2025"] = row["total_articles"]
        if row["total_citations"] is not None:
            rec["jcr_total_citations_2025"] = row["total_citations"]

        pub_history = output_history_for(rec, annual_outputs)
        if pub_history:
            rec["publication_history"] = pub_history
        matched += 1

    pub_attached = 0
    if annual_outputs:
        for rec in journals:
            if rec.get("publication_history"):
                pub_attached += 1
                continue
            pub_history = output_history_for(rec, annual_outputs)
            if pub_history:
                rec["publication_history"] = pub_history
                pub_attached += 1
    return matched, pub_attached


def compact_dict(value, keys: tuple[str, ...]):
    if not isinstance(value, dict):
        return value
    return {k: value[k] for k in keys if value.get(k) not in (None, "", [], {})}


def compact_journal_for_light(rec: dict) -> dict:
    base_keys = (
        "name", "cn_name", "en_name", "abbr20", "slug", "issn", "eissn",
        "publisher", "country", "indices", "wos_categories", "esi_category",
        "if_2024", "if_2025", "if_latest", "if_latest_year", "if_quartile",
        "jif_without_self_cites_2025", "self_citation_rate_2025",
        "jcr_year", "jcr_release_year", "cas_zone", "cas_top", "cas_major_cn",
        "flagship", "nature_index", "free", "pubmed", "pmc", "medline",
        "under_review", "on_hold", "citic_warning", "ccf", "ft50", "utd24",
    )
    out = {k: rec[k] for k in base_keys if rec.get(k) not in (None, "", [], {})}
    if rec.get("scopus") and rec.get("scopus", {}).get("active", True) is not False:
        out["scopus"] = {"active": True}
    for key, keys in {
        "doaj": ("lic", "license", "review_weeks"),
        "oaj": ("partition", "position"),
        "cas_xr": ("zone", "top"),
        "cscd": ("database",),
        "cstpcd": ("kind", "year"),
        "scd": ("year", "category", "newly_added"),
        "abdc": ("rating",),
        "abs": ("rating",),
        "fms": ("tier", "type"),
        "retraction": ("retractions_total", "rate_per_1000_5y", "rate_per_1000_10y"),
    }.items():
        value = compact_dict(rec.get(key), keys)
        if value not in (None, "", [], {}):
            out[key] = value
    warning = rec.get("warning")
    if warning:
        if isinstance(warning, list):
            slim = [compact_dict(x, ("year", "level")) for x in warning if isinstance(x, dict)]
            out["warning"] = slim or True
        elif isinstance(warning, dict):
            out["warning"] = compact_dict(warning, ("year", "level")) or True
        else:
            out["warning"] = True
    return out


def write_light_index(journals: list[dict]) -> int:
    light = [compact_journal_for_light(rec) for rec in journals]
    write_json_gz(LIGHT_GZ, light)
    # The homepage/detail bootstrap consumes v2; keep it in sync so the
    # latest JCR fields are available before the full bundle upgrade.
    write_json_gz(LIGHT_V2_GZ, light)
    return len(light)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=None)
    args = parser.parse_args()

    workbook = args.workbook or first_existing(DEFAULT_WORKBOOKS)
    if workbook is None:
        raise SystemExit("Latest JCR workbook not found")
    annual_path = first_existing(ANNUAL_OUTPUT_CANDIDATES)
    annual_outputs = read_json_gz(annual_path) if annual_path else {}

    journals = read_json_gz(JOURNALS_GZ)
    latest_rows = read_latest_jcr(workbook)
    matched, pub_attached = apply_updates(journals, latest_rows, annual_outputs)
    free_attached = attach_free_flags(journals)
    write_json_bundle(JOURNALS_JSON, JOURNALS_GZ, journals)

    light_records = write_light_index(journals)

    meta_path = DATA_DIR / "meta.json"
    if meta_path.exists():
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
    else:
        meta = {}
    meta["jcr_latest_metric_year"] = 2025
    meta["jcr_latest_release_year"] = 2026
    meta["with_if_2025"] = sum(rec.get("if_2025") not in (None, "") for rec in journals)
    meta["with_publication_history"] = pub_attached
    meta["with_free_to_publish"] = free_attached
    meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Workbook: {workbook}")
    print(f"Latest JCR rows: {len(latest_rows)}")
    print(f"Matched journals: {matched}")
    print(f"Publication histories attached: {pub_attached}")
    print(f"Free-to-publish flags attached: {free_attached}")
    print(f"Light index records written: {light_records}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
